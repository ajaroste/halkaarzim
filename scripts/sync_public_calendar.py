from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Iterable

import requests
from openpyxl import load_workbook

DEFAULT_PUBLIC_CALENDAR_URL = (
    "https://api.halkaarz.com.tr/api/companies/export?format=xlsx&group=first&limit=700"
)
SOURCE_TITLE = "Kamuya açık halka arz takvimi dışa aktarımı"
SOURCE_KIND = "İkincil kamuya açık veri"

STATUS_LABELS = {
    "active": "Talep topluyor",
    "upcoming": "Yaklaşan",
    "approved": "SPK onaylı",
    "completed": "Arzı tamamlandı",
    "listed": "İşlem görüyor",
    "delayed": "Ertelendi",
    "draft": "Taslak",
}

TR_MONTHS = {
    "ocak": 1,
    "subat": 2,
    "mart": 3,
    "nisan": 4,
    "mayis": 5,
    "haziran": 6,
    "temmuz": 7,
    "agustos": 8,
    "eylul": 9,
    "ekim": 10,
    "kasim": 11,
    "aralik": 12,
}
TR_MONTH_NAMES = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "company": (
        "sirket", "sirket adi", "sirket unvani", "firma", "firma adi", "company", "company name",
        "halka arz sirketi", "halka arz olan sirket",
    ),
    "ticker": (
        "kod", "borsa kodu", "hisse kodu", "islem kodu", "sembol", "ticker", "symbol", "stock code",
    ),
    "dates": (
        "tarih", "tarihler", "halka arz tarihi", "halka arz tarihleri", "talep tarihi", "talep tarihleri",
        "talep toplama tarihi", "talep toplama tarihleri", "dates", "demand dates", "collection dates",
    ),
    "collectionStart": (
        "talep baslangic", "talep baslangic tarihi", "talep toplama baslangic", "talep toplama baslangic tarihi",
        "baslangic tarihi", "collection start", "demand start", "start date",
    ),
    "collectionEnd": (
        "talep bitis", "talep bitis tarihi", "talep toplama bitis", "talep toplama bitis tarihi", "bitis tarihi",
        "collection end", "demand end", "end date",
    ),
    "firstTradeDate": (
        "ilk islem tarihi", "borsada islem tarihi", "islem gorme tarihi", "islem tarihi", "first trade date",
        "listing date",
    ),
    "participantCount": (
        "katilimci", "katilimci sayisi", "yatirimci sayisi", "participant count", "investor count",
    ),
    "distribution": (
        "dagitim", "dagitim yontemi", "talep toplama yontemi", "distribution", "distribution method",
    ),
    "intermediary": (
        "araci kurum", "araci kurumlar", "konsorsiyum lideri", "lider araci kurum", "intermediary", "broker",
    ),
    "market": (
        "pazar", "borsa pazari", "islem gorecegi pazar", "market", "market segment",
    ),
    "publicFloat": (
        "halka aciklik", "halka aciklik orani", "public float", "free float",
    ),
    "offerSize": (
        "halka arz buyuklugu", "arz buyuklugu", "offer size", "ipo size",
    ),
    "price": (
        "fiyat", "halka arz fiyati", "pay basi fiyat", "price", "offer price",
    ),
    "status": (
        "durum", "surec", "surec durumu", "halka arz durumu", "status", "state",
    ),
}

DYNAMIC_FIELDS = {
    "ticker", "collectionStart", "collectionEnd", "firstTradeDate", "participantCount", "distribution",
    "intermediary", "market", "publicFloat", "offerSize", "price",
}


class PublicCalendarSyncError(RuntimeError):
    pass


@dataclass(frozen=True)
class SyncReport:
    source_url: str
    source_rows: int
    matched_rows: int
    changed_items: int
    unmatched_companies: tuple[str, ...]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("ı", "i")
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def normalize_company_name(value: Any) -> str:
    tokens = normalize_text(value).split()
    ignored = {
        "a", "s", "as", "anonim", "sirket", "sirketi", "san", "sanayi", "tic", "ticaret", "ve",
        "limited", "ltd", "sti", "st", "inc", "corp", "corporation",
    }
    return " ".join(token for token in tokens if token not in ignored)


def _alias_score(header: str, aliases: Iterable[str]) -> int:
    if not header:
        return 0
    header_tokens = set(header.split())
    best = 0
    for alias in aliases:
        norm = normalize_text(alias)
        if header == norm:
            best = max(best, 100)
            continue
        alias_tokens = set(norm.split())
        if norm and (norm in header or header in norm):
            best = max(best, 85)
        if alias_tokens and alias_tokens.issubset(header_tokens):
            best = max(best, 80 + min(len(alias_tokens), 4))
    return best


def map_headers(headers: list[Any]) -> dict[str, int]:
    normalized = [normalize_text(value) for value in headers]
    mapping: dict[str, int] = {}
    used: set[int] = set()
    for field, aliases in FIELD_ALIASES.items():
        ranked = sorted(
            ((_alias_score(header, aliases), index) for index, header in enumerate(normalized) if index not in used),
            reverse=True,
        )
        if ranked and ranked[0][0] >= 80:
            mapping[field] = ranked[0][1]
            used.add(ranked[0][1])
    return mapping


def _detect_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    best: tuple[int, int, dict[str, int]] | None = None
    for index, row in enumerate(rows[:15]):
        mapping = map_headers(row)
        if "company" not in mapping:
            continue
        dynamic_count = sum(field in mapping for field in DYNAMIC_FIELDS | {"dates", "status"})
        score = 10 + dynamic_count
        if best is None or score > best[0]:
            best = (score, index, mapping)
    if best is None or best[0] < 11:
        preview = [[str(cell)[:80] if cell is not None else "" for cell in row[:12]] for row in rows[:8]]
        raise PublicCalendarSyncError(f"Takvim dosyasında desteklenen başlıklar bulunamadı. İlk satırlar: {preview}")
    return best[1], best[2]


def _cell(row: list[Any], mapping: dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    return row[index] if index is not None and index < len(row) else None


def parse_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover
        raise PublicCalendarSyncError(f"Takvim XLSX dosyası açılamadı: {exc}") from exc

    candidates: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        try:
            header_index, mapping = _detect_header_row(rows)
        except PublicCalendarSyncError:
            continue
        for row in rows[header_index + 1 :]:
            company = _cell(row, mapping, "company")
            if not company or not normalize_company_name(company):
                continue
            candidates.append({field: _cell(row, mapping, field) for field in mapping})
    if not candidates:
        raise PublicCalendarSyncError("Takvim XLSX dosyasında şirket satırı bulunamadı")
    return candidates


def _extract_json_rows(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for key in ("items", "data", "results", "companies", "rows"):
            if key in payload:
                rows = _extract_json_rows(payload[key])
                if rows:
                    return rows
    return []


def parse_json_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        payload = json.loads(content.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise PublicCalendarSyncError(f"Takvim JSON yanıtı okunamadı: {exc}") from exc
    raw_rows = _extract_json_rows(payload)
    if not raw_rows:
        raise PublicCalendarSyncError("Takvim JSON yanıtında şirket satırı bulunamadı")

    all_keys: list[str] = []
    for row in raw_rows[:20]:
        for key in row:
            if key not in all_keys:
                all_keys.append(key)
    mapping = map_headers(all_keys)
    if "company" not in mapping:
        raise PublicCalendarSyncError(f"Takvim JSON alanları eşleştirilemedi: {all_keys[:40]}")

    key_for_field = {field: all_keys[index] for field, index in mapping.items()}
    return [
        {field: row.get(key) for field, key in key_for_field.items()}
        for row in raw_rows
        if row.get(key_for_field["company"])
    ]


def fetch_public_rows(url: str = DEFAULT_PUBLIC_CALENDAR_URL, timeout: int = 35) -> list[dict[str, Any]]:
    headers = {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/json;q=0.9, */*;q=0.5",
        "User-Agent": "HalkaArzimDataSync/1.0 (+https://halkaarzim.vercel.app)",
    }
    try:
        response = requests.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise PublicCalendarSyncError(f"Takvim kaynağına erişilemedi: {exc}") from exc

    content_type = response.headers.get("content-type", "").lower()
    content = response.content
    if not content:
        raise PublicCalendarSyncError("Takvim kaynağı boş yanıt döndürdü")
    if "json" in content_type or content.lstrip().startswith((b"{", b"[")):
        return parse_json_rows(content)
    if content.startswith(b"PK") or "spreadsheet" in content_type or "excel" in content_type:
        return parse_xlsx_rows(content)
    raise PublicCalendarSyncError(f"Takvim kaynağının içerik tipi desteklenmiyor: {content_type or 'bilinmiyor'}")


def parse_date_value(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def parse_collection_dates(value: Any) -> tuple[date | None, date | None]:
    if isinstance(value, datetime):
        return value.date(), value.date()
    if isinstance(value, date):
        return value, value
    text = str(value or "").strip()
    if not text:
        return None, None
    normalized = normalize_text(text)

    explicit: list[date] = []
    for day, month, year in re.findall(r"\b(\d{1,2})[./-](\d{1,2})[./-](20\d{2})\b", text):
        try:
            explicit.append(date(int(year), int(month), int(day)))
        except ValueError:
            pass
    if explicit:
        return min(explicit), max(explicit)

    single = parse_date_value(value)
    if single:
        return single, single

    month_pattern = "|".join(TR_MONTHS)
    month_match = re.search(rf"\b({month_pattern})\b", normalized)
    year_match = re.search(r"\b(20\d{2})\b", normalized)

    if month_match and year_match:
        prefix = normalized[: month_match.start()]
        days = [int(value) for value in re.findall(r"\b(\d{1,2})\b", prefix) if 1 <= int(value) <= 31]
        if len(days) > 1:
            parsed: list[date] = []
            for day in days:
                try:
                    parsed.append(date(int(year_match.group(1)), TR_MONTHS[month_match.group(1)], day))
                except ValueError:
                    pass
            if parsed:
                return min(parsed), max(parsed)

    mentions = list(re.finditer(rf"\b(\d{{1,2}})\s+({month_pattern})(?:\s+(20\d{{2}}))?", normalized))
    if mentions:
        default_year = int(year_match.group(1)) if year_match else date.today().year
        dates: list[date] = []
        for match in mentions:
            year = int(match.group(3)) if match.group(3) else default_year
            try:
                dates.append(date(year, TR_MONTHS[match.group(2)], int(match.group(1))))
            except ValueError:
                pass
        if dates:
            return min(dates), max(dates)

    if month_match and year_match:
        prefix = normalized[: month_match.start()]
        days = [int(value) for value in re.findall(r"\b(\d{1,2})\b", prefix) if 1 <= int(value) <= 31]
        parsed: list[date] = []
        for day in days:
            try:
                parsed.append(date(int(year_match.group(1)), TR_MONTHS[month_match.group(1)], day))
            except ValueError:
                pass
        if parsed:
            return min(parsed), max(parsed)
    return None, None


def format_collection_dates(start: date | None, end: date | None) -> str | None:
    if not start and not end:
        return None
    start = start or end
    end = end or start
    assert start is not None and end is not None
    if end < start:
        start, end = end, start
    span = (end - start).days
    if start.year == end.year and start.month == end.month:
        if span <= 5:
            days = "-".join(str((start + timedelta(days=offset)).day) for offset in range(span + 1))
            return f"{days} {TR_MONTH_NAMES[start.month]} {start.year}"
        return f"{start.day}-{end.day} {TR_MONTH_NAMES[start.month]} {start.year}"
    return f"{start.day} {TR_MONTH_NAMES[start.month]} {start.year} - {end.day} {TR_MONTH_NAMES[end.month]} {end.year}"


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    digits = re.sub(r"[^0-9]", "", str(value))
    return int(digits) if digits else None


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = re.sub(r"[^0-9,.-]", "", str(value).strip())
    if not text:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        text = text.replace(".", "")
    try:
        return float(text)
    except ValueError:
        return None


def _clean_string(value: Any) -> str | None:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text or normalize_text(text) in {"null", "none", "yok", "-"}:
        return None
    return text


def _clean_ticker(value: Any) -> str | None:
    text = _clean_string(value)
    if not text:
        return None
    ticker = text.upper().replace("BIST:", "").strip()
    ticker = re.sub(r"\.(H|E)$", "", ticker)
    ticker = re.sub(r"[^A-Z0-9]", "", ticker)
    return ticker or None


def canonicalize_source_row(row: dict[str, Any]) -> dict[str, Any]:
    company = _clean_string(row.get("company"))
    start = parse_date_value(row.get("collectionStart"))
    end = parse_date_value(row.get("collectionEnd"))
    if not start or not end:
        combined_start, combined_end = parse_collection_dates(row.get("dates"))
        start = start or combined_start
        end = end or combined_end

    first_trade = parse_date_value(row.get("firstTradeDate"))
    result: dict[str, Any] = {
        "company": company,
        "ticker": _clean_ticker(row.get("ticker")),
        "collectionStart": start.isoformat() if start else None,
        "collectionEnd": end.isoformat() if end else None,
        "firstTradeDate": first_trade.isoformat() if first_trade else None,
        "participantCount": _to_int(row.get("participantCount")),
        "distribution": _clean_string(row.get("distribution")),
        "intermediary": _clean_string(row.get("intermediary")),
        "market": _clean_string(row.get("market")),
        "publicFloat": _to_float(row.get("publicFloat")),
        "offerSize": _to_float(row.get("offerSize")),
        "price": _to_float(row.get("price")),
        "sourceStatus": _clean_string(row.get("status")),
    }
    dates_label = format_collection_dates(start, end)
    if dates_label:
        result["dates"] = dates_label
    return result


def _match_score(existing_key: str, source_key: str) -> float:
    if existing_key == source_key:
        return 1.0
    existing_tokens = set(existing_key.split())
    source_tokens = set(source_key.split())
    if not existing_tokens or not source_tokens:
        return 0.0
    overlap = len(existing_tokens & source_tokens) / max(len(existing_tokens), len(source_tokens))
    prefix = 1.0 if (existing_key.startswith(source_key) or source_key.startswith(existing_key)) else 0.0
    return max(overlap, prefix * 0.92)


def _find_source_row(item: dict[str, Any], source_rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    target = normalize_company_name(item.get("company"))
    if not target:
        return None
    exact = [row for row in source_rows if normalize_company_name(row.get("company")) == target]
    if exact:
        return exact[0]
    ranked = sorted(
        ((_match_score(target, normalize_company_name(row.get("company"))), row) for row in source_rows),
        key=lambda pair: pair[0],
        reverse=True,
    )
    return ranked[0][1] if ranked and ranked[0][0] >= 0.88 else None


def _status_from_source_text(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    if "ertelen" in text or "postpon" in text:
        return "delayed"
    if "islem gor" in text or "listed" in text or "trading" in text:
        return "listed"
    if "talep" in text or "active" in text or "collect" in text:
        return "active"
    if "yaklas" in text or "upcoming" in text:
        return "upcoming"
    if "tamam" in text or "completed" in text:
        return "completed"
    if "onay" in text or "approved" in text:
        return "approved"
    return None


def merge_public_calendar(
    items: list[dict[str, Any]], source_rows: list[dict[str, Any]], source_url: str = DEFAULT_PUBLIC_CALENDAR_URL
) -> tuple[list[dict[str, Any]], SyncReport]:
    canonical_rows = [canonicalize_source_row(row) for row in source_rows]
    canonical_rows = [row for row in canonical_rows if row.get("company")]
    output: list[dict[str, Any]] = []
    matched_source_ids: set[int] = set()
    changed_items = 0
    matched_rows = 0

    for original in items:
        item = dict(original)
        row = _find_source_row(item, canonical_rows)
        if not row:
            output.append(item)
            continue
        matched_rows += 1
        matched_source_ids.add(id(row))
        before = json.dumps(item, ensure_ascii=False, sort_keys=True, default=str)

        for field in (
            "ticker", "collectionStart", "collectionEnd", "firstTradeDate", "participantCount", "distribution",
            "intermediary", "market", "publicFloat", "offerSize",
        ):
            value = row.get(field)
            if value not in (None, ""):
                item[field] = value

        # SPK is authoritative for price. The secondary source only fills a missing price.
        if (not item.get("price") or float(item.get("price") or 0) <= 0) and row.get("price"):
            item["price"] = row["price"]

        if row.get("dates"):
            item["dates"] = row["dates"]

        fallback_status = _status_from_source_text(row.get("sourceStatus"))
        if fallback_status == "delayed":
            item["postponed"] = True
        elif fallback_status and not (item.get("collectionStart") and item.get("collectionEnd")) and not item.get("firstTradeDate"):
            item["status"] = fallback_status
            item["statusLabel"] = STATUS_LABELS[fallback_status]

        after_without_source = json.dumps(item, ensure_ascii=False, sort_keys=True, default=str)
        if after_without_source != before:
            sources = list(item.get("sources") or [])
            if not any(source.get("url") == source_url for source in sources):
                sources.append({
                    "title": SOURCE_TITLE,
                    "page": "Takvim ve sonuç alanları",
                    "kind": SOURCE_KIND,
                    "url": source_url,
                })
            item["sources"] = sources
            item["sourceUpdatedAt"] = datetime.now().astimezone().isoformat()
            changed_items += 1
        output.append(item)

    unmatched = tuple(str(row.get("company")) for row in canonical_rows if id(row) not in matched_source_ids)
    return output, SyncReport(
        source_url=source_url,
        source_rows=len(canonical_rows),
        matched_rows=matched_rows,
        changed_items=changed_items,
        unmatched_companies=unmatched[:25],
    )
