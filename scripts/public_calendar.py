from __future__ import annotations

import os
import re
import time
import unicodedata
import uuid
from copy import deepcopy
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Iterable
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

from scripts.enrich_ipo_data import clean_ticker, compute_status, normalize_company_name

DEFAULT_EXPORT_URL = "https://api.halkaarz.com.tr/api/companies/export?format=xlsx&group=first&limit=700"
PUBLIC_SOURCE_TITLE = "HalkaArz.com.tr halka arz takvimi"
PUBLIC_SOURCE_KIND = "İkincil kamuya açık veri"
NAMESPACE = uuid.UUID("21c44ed5-9308-4b0d-8f62-78a421205aec")

MONTHS = {
    "ocak": 1,
    "subat": 2,
    "mart": 3,
    "nisan": 4,
    "mayis": 5,
    "haziran": 6,
    "temmuz": 7,
    "agustos": 8,
    "eylul": 9,
    "ekim": 10,
    "kasim": 11,
    "aralik": 12,
}
MONTH_NAMES = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "company": (
        "sirket", "sirket adi", "sirket unvani", "firma", "firma adi", "unvan", "company", "company name",
    ),
    "ticker": (
        "kod", "bist kodu", "borsa kodu", "hisse kodu", "islem kodu", "sembol", "ticker", "symbol",
    ),
    "status": ("durum", "halka arz durumu", "arz durumu", "surec", "surec durumu", "status"),
    "dates": (
        "halka arz tarihi", "halka arz tarihleri", "talep toplama tarihi", "talep toplama tarihleri",
        "talep tarihleri", "talep tarihi", "talep toplama", "tarihler", "tarih",
    ),
    "collectionStart": (
        "talep baslangic", "talep baslangic tarihi", "talep toplama baslangic",
        "talep toplama baslangic tarihi", "baslangic tarihi", "collection start", "demand start",
    ),
    "collectionEnd": (
        "talep bitis", "talep bitis tarihi", "talep toplama bitis", "talep toplama bitis tarihi",
        "bitis tarihi", "collection end", "demand end",
    ),
    "firstTradeDate": (
        "ilk islem tarihi", "borsada islem tarihi", "islem baslangic tarihi", "islem gorme tarihi",
        "first trade date", "listing date",
    ),
    "price": ("halka arz fiyati", "arz fiyati", "pay fiyati", "pay basi fiyat", "fiyat", "price"),
    "lotCount": (
        "pay lot", "pay lot adedi", "pay adedi", "lot", "lot sayisi", "arz edilen pay", "pay sayisi", "pay",
    ),
    "distribution": ("dagitim yontemi", "talep toplama yontemi", "dagitim", "distribution"),
    "intermediary": (
        "araci kurum", "araci kurumlar", "konsorsiyum lideri", "lider araci kurum", "intermediary", "broker",
    ),
    "sector": ("sektor", "sector"),
    "market": ("islem gorecegi pazar", "borsa pazari", "pazar yeri", "pazar", "market segment", "market"),
    "participantCount": (
        "katilimci", "katilimci sayisi", "yatirimci sayisi", "participants", "participant count", "investor count",
    ),
    "offerSize": ("halka arz buyuklugu", "arz buyuklugu", "buyukluk", "offer size", "ipo size"),
}

STATUS_LABELS = {
    "active": "Talep topluyor",
    "upcoming": "Yaklaşan",
    "approved": "SPK onaylı",
    "completed": "Arzı tamamlandı",
    "listed": "İşlem görüyor",
    "delayed": "Ertelendi",
    "draft": "Taslak",
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ı", "i")
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _header_score(header: str, alias: str) -> int:
    if not header or not alias:
        return 0
    if header == alias:
        return 100
    if header.startswith(alias + " "):
        return 92
    alias_tokens = alias.split()
    header_tokens = set(header.split())
    if len(alias_tokens) >= 2 and set(alias_tokens).issubset(header_tokens):
        return 80 + min(len(alias_tokens), 8)
    return 0


def _header_field(value: Any) -> str | None:
    normalized = normalize_header(value)
    if not normalized:
        return None
    best_field: str | None = None
    best_score = 0
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            score = _header_score(normalized, normalize_header(alias))
            if score > best_score:
                best_field = field
                best_score = score
    return best_field if best_score >= 80 else None


def _best_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]]:
    best_index = -1
    best_map: dict[int, str] = {}
    for index, row in enumerate(rows[:15]):
        mapping: dict[int, str] = {}
        for column, value in enumerate(row):
            field = _header_field(value)
            if field and field not in mapping.values():
                mapping[column] = field
        if "company" in mapping.values() and len(mapping) > len(best_map):
            best_index, best_map = index, mapping
    if best_index < 0 or len(best_map) < 2:
        seen = [normalize_header(value) for row in rows[:8] for value in row if value]
        raise ValueError(f"Halka arz Excel başlıkları tanınamadı: {seen[:30]}")
    return best_index, best_map


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().lower()
    if not text or normalize_header(text) in {"hazirlaniyor", "aciklanmadi", "yok"}:
        return None
    multiplier = 1.0
    if "milyar" in text:
        multiplier = 1_000_000_000.0
    elif "milyon" in text:
        multiplier = 1_000_000.0
    elif "bin" in text:
        multiplier = 1_000.0
    cleaned = re.sub(r"[^0-9,.-]", "", text)
    if not cleaned:
        return None
    if "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif cleaned.count(".") > 1 or ("." in cleaned and len(cleaned.rsplit(".", 1)[1]) == 3):
        cleaned = cleaned.replace(".", "")
    try:
        return float(cleaned) * multiplier
    except ValueError:
        return None


def parse_turkish_date_range(value: Any) -> tuple[str | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        iso = value.date().isoformat()
        return iso, iso
    if isinstance(value, date):
        iso = value.isoformat()
        return iso, iso
    text = _as_text(value)
    if not text:
        return None, None
    normalized = normalize_header(text)
    if any(token in normalized for token in ("hazirlaniyor", "aciklanmadi", "belli olmadi", "bekleniyor")):
        return None, None

    iso_dates = re.findall(r"\b(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})\b", text)
    if iso_dates:
        parsed = [date(int(y), int(m), int(d)) for y, m, d in iso_dates]
        return min(parsed).isoformat(), max(parsed).isoformat()

    tr_numeric_dates = re.findall(r"\b(\d{1,2})[./-](\d{1,2})[./-](20\d{2})\b", text)
    if tr_numeric_dates:
        parsed = [date(int(y), int(m), int(d)) for d, m, y in tr_numeric_dates]
        return min(parsed).isoformat(), max(parsed).isoformat()

    year_match = re.search(r"\b(20\d{2})\b", text)
    if not year_match:
        return None, None
    year = int(year_match.group(1))
    month_pattern = re.compile(
        r"\b(Ocak|Şubat|Subat|Mart|Nisan|Mayıs|Mayis|Haziran|Temmuz|Ağustos|Agustos|Eylül|Eylul|Ekim|Kasım|Kasim|Aralık|Aralik)\b",
        re.I,
    )
    matches = list(month_pattern.finditer(text))
    parsed: list[date] = []
    segment_start = 0
    for match in matches:
        month_name = normalize_header(match.group(1))
        month = MONTHS.get(month_name)
        if not month:
            continue
        segment = text[segment_start:match.start()]
        segment = re.sub(r"\b20\d{2}\b", "", segment)
        # Times such as 09:00-17:00 appear after the month in the current export,
        # so only the segment before the month contributes day numbers.
        days = [int(x) for x in re.findall(r"(?<!\d)(\d{1,2})(?!\d)", segment) if 1 <= int(x) <= 31]
        for day in days:
            try:
                parsed.append(date(year, month, day))
            except ValueError:
                pass
        segment_start = match.end()
    if not parsed:
        return None, None
    return min(parsed).isoformat(), max(parsed).isoformat()


def _format_collection_dates(start: str | None, end: str | None) -> str | None:
    if not start and not end:
        return None
    try:
        start_date = date.fromisoformat((start or end or "")[:10])
        end_date = date.fromisoformat((end or start or "")[:10])
    except ValueError:
        return None
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    span = (end_date - start_date).days
    if start_date.year == end_date.year and start_date.month == end_date.month:
        if span <= 5:
            days = "-".join(str((start_date + timedelta(days=offset)).day) for offset in range(span + 1))
            return f"{days} {MONTH_NAMES[start_date.month]} {start_date.year}"
        return f"{start_date.day}-{end_date.day} {MONTH_NAMES[start_date.month]} {start_date.year}"
    return (
        f"{start_date.day} {MONTH_NAMES[start_date.month]} {start_date.year} - "
        f"{end_date.day} {MONTH_NAMES[end_date.month]} {end_date.year}"
    )


def _status_from_text(value: Any) -> str | None:
    normalized = normalize_header(value)
    if not normalized:
        return None
    if "ertel" in normalized or "iptal" in normalized:
        return "delayed"
    if "islem gor" in normalized or "islemde" in normalized or "borsada" in normalized:
        return "listed"
    if "tamam" in normalized:
        return "completed"
    if "talep" in normalized and ("topla" in normalized or "acik" in normalized):
        return "active"
    if "yaklas" in normalized or "tarihi acik" in normalized:
        return "upcoming"
    if "taslak" in normalized or "basvuru" in normalized:
        return "draft"
    if "spk" in normalized or "onay" in normalized:
        return "approved"
    return None


def _date_candidate_from_unmapped(row: tuple[Any, ...], mapped_columns: set[int]) -> str | None:
    for column, value in enumerate(row):
        if column in mapped_columns:
            continue
        text = _as_text(value)
        if not text or not re.search(r"\b20\d{2}\b", text):
            continue
        start, end = parse_turkish_date_range(text)
        if start and end:
            return text
    return None


def _valid_market(value: str | None) -> bool:
    text = normalize_header(value)
    if not text:
        return False
    return any(token in text for token in ("pazar", "market", "yildiz", "ana pazar", "alt pazar"))


def parse_export(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_index, mapping = _best_header_row(rows)
    records: list[dict[str, Any]] = []
    mapped_columns = set(mapping)

    for row in rows[header_index + 1:]:
        record: dict[str, Any] = {}
        for column, field in mapping.items():
            if column < len(row):
                record[field] = row[column]
        company = _as_text(record.get("company"))
        if not company:
            continue

        output: dict[str, Any] = {"company": company}
        ticker = clean_ticker(_as_text(record.get("ticker")))
        if ticker:
            output["ticker"] = ticker
        status = _status_from_text(record.get("status"))
        if status:
            output["status"] = status

        dates = _as_text(record.get("dates")) or _date_candidate_from_unmapped(row, mapped_columns)
        start = _as_text(record.get("collectionStart"))
        end = _as_text(record.get("collectionEnd"))
        if start:
            parsed_start, _ = parse_turkish_date_range(start)
            start = parsed_start or start[:10]
        if end:
            _, parsed_end = parse_turkish_date_range(end)
            end = parsed_end or end[:10]
        if (not start or not end) and dates:
            parsed_start, parsed_end = parse_turkish_date_range(dates)
            start = start or parsed_start
            end = end or parsed_end
        if start:
            output["collectionStart"] = start
        if end:
            output["collectionEnd"] = end
        clean_dates = _format_collection_dates(start, end)
        if clean_dates:
            output["dates"] = clean_dates
        elif dates and normalize_header(dates) not in {"hazirlaniyor", "aciklanmadi", "bekleniyor"}:
            output["dates"] = dates

        trade = _as_text(record.get("firstTradeDate"))
        if trade:
            parsed_trade, _ = parse_turkish_date_range(trade)
            output["firstTradeDate"] = parsed_trade or trade[:10]

        for field in ("distribution", "intermediary", "sector"):
            text = _as_text(record.get(field))
            if text and normalize_header(text) not in {"hazirlaniyor", "aciklanmadi"}:
                output[field] = text
        market = _as_text(record.get("market"))
        if market and _valid_market(market):
            output["market"] = market

        for field in ("price", "lotCount", "participantCount", "offerSize"):
            number = _number(record.get(field))
            if number is not None:
                output[field] = int(number) if number.is_integer() else number
        records.append(output)

    if not records:
        raise ValueError("Halka arz Excel dosyasında şirket kaydı bulunamadı")
    return records


def fetch_public_export(url: str | None = None, timeout: int = 30) -> tuple[list[dict[str, Any]], str]:
    url = url or os.getenv("IPO_PUBLIC_EXPORT_URL", DEFAULT_EXPORT_URL)
    headers = {
        "User-Agent": "HalkaArzimDataSync/1.0 (+https://halkaarzim.vercel.app)",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream;q=0.9,*/*;q=0.5",
    }
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            response = requests.get(url, headers=headers, timeout=timeout)
            response.raise_for_status()
            if not response.content.startswith(b"PK"):
                raise ValueError(f"Beklenmeyen Excel yanıtı: {response.headers.get('content-type', 'bilinmiyor')}")
            return parse_export(response.content), url
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt < 2:
                time.sleep(2 ** attempt)
    raise RuntimeError(f"Halka arz takvim kaynağı alınamadı: {last_error}") from last_error


def _source(url: str) -> dict[str, str]:
    return {
        "title": PUBLIC_SOURCE_TITLE,
        "page": "İlk halka arzlar Excel dışa aktarımı",
        "kind": PUBLIC_SOURCE_KIND,
        "url": url,
    }


def _is_missing_number(value: Any) -> bool:
    try:
        return value is None or float(value) <= 0
    except (TypeError, ValueError):
        return True


def _merge_record(item: dict[str, Any], record: dict[str, Any], source_url: str, now_iso: str) -> tuple[dict[str, Any], bool]:
    merged = deepcopy(item)
    changed = False

    for field in (
        "ticker", "collectionStart", "collectionEnd", "firstTradeDate", "dates", "distribution",
        "intermediary", "market", "participantCount", "offerSize",
    ):
        value = record.get(field)
        if value not in (None, "") and merged.get(field) != value:
            merged[field] = value
            changed = True

    # SPK-derived core values remain authoritative; the calendar may only fill gaps.
    for field in ("price", "lotCount"):
        value = record.get(field)
        if value not in (None, "") and _is_missing_number(merged.get(field)):
            merged[field] = value
            changed = True
    if (not merged.get("sector") or merged.get("sector") == "Diğer") and record.get("sector"):
        merged["sector"] = record["sector"]
        changed = True

    if record.get("status") and merged.get("status") != record["status"]:
        merged["status"] = record["status"]
        changed = True
    computed = compute_status(merged)
    if merged.get("status") != computed:
        merged["status"] = computed
        changed = True
    label = STATUS_LABELS.get(computed, merged.get("statusLabel") or "SPK onaylı")
    if merged.get("statusLabel") != label:
        merged["statusLabel"] = label
        changed = True

    sources = list(merged.get("sources") or [])
    public_source = _source(source_url)
    if not any(source.get("url") == source_url for source in sources):
        sources.append(public_source)
        merged["sources"] = sources
        changed = True
    if changed:
        merged["sourceUpdatedAt"] = now_iso
    return merged, changed


def _company_similarity(left: str, right: str) -> float:
    if left == right:
        return 1.0
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    if not left_tokens or not right_tokens:
        return 0.0
    overlap = len(left_tokens & right_tokens) / max(len(left_tokens), len(right_tokens))
    prefix = 0.92 if left.startswith(right) or right.startswith(left) else 0.0
    return max(overlap, prefix)


def _find_existing_index(key: str, by_name: dict[str, int]) -> int | None:
    if key in by_name:
        return by_name[key]
    best_score = 0.0
    best_index: int | None = None
    for existing_key, index in by_name.items():
        score = _company_similarity(key, existing_key)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index if best_score >= 0.88 else None


def _should_add_new(record: dict[str, Any], today: date) -> bool:
    status = record.get("status")
    if status in {"active", "upcoming", "approved"}:
        return True
    start = record.get("collectionStart")
    if start:
        try:
            start_date = date.fromisoformat(str(start)[:10])
            return today - timedelta(days=60) <= start_date <= today + timedelta(days=120)
        except ValueError:
            pass
    return False


def _new_item(record: dict[str, Any], source_url: str, now_iso: str) -> dict[str, Any]:
    company = str(record["company"])
    stable_id = str(uuid.uuid5(NAMESPACE, normalize_company_name(company)))
    item: dict[str, Any] = {
        "id": stable_id,
        "slug": re.sub(r"[^a-z0-9]+", "-", normalize_header(company)).strip("-"),
        "ticker": record.get("ticker"),
        "company": company,
        "sector": record.get("sector") or "Diğer",
        "status": record.get("status") or "approved",
        "statusLabel": STATUS_LABELS.get(record.get("status") or "approved", "SPK onaylı"),
        "price": record.get("price") or 0,
        "dates": record.get("dates") or "Talep toplama tarihleri henüz açıklanmadı",
        "collectionStart": record.get("collectionStart"),
        "collectionEnd": record.get("collectionEnd"),
        "firstTradeDate": record.get("firstTradeDate"),
        "participantCount": record.get("participantCount") or 0,
        "offerSize": record.get("offerSize") or 0,
        "intermediary": record.get("intermediary"),
        "publicFloat": 0,
        "market": record.get("market") or "Henüz açıklanmadı",
        "priceStability": "Henüz açıklanmadı",
        "valuationDiscount": 0,
        "allocationText": "",
        "dataCompleteness": 35,
        "dataNotes": [
            "Kayıt otomatik halka arz takvim senkronizasyonuyla eklendi; resmî doküman alanları tamamlandıkça güncellenir."
        ],
        "lotCount": record.get("lotCount") or 0,
        "maxLotCount": record.get("lotCount") or 0,
        "retailLots": 0,
        "distribution": record.get("distribution") or "Henüz açıklanmadı",
        "aiScore": 50,
        "risk": "Belirsiz",
        "aiSummary": "Otomatik takvim kaydıdır; resmî doküman alanları tamamlanmadan tam analiz üretilmez.",
        "aiProvider": "rules-v1",
        "reportVersion": "otomatik-on-analiz-1",
        "reportDate": datetime.now(ZoneInfo("Europe/Istanbul")).date().isoformat(),
        "humanReviewed": False,
        "analysisStatus": "preliminary",
        "analysisScope": "Takvim, durum ve temel halka arz alanları",
        "capitalBefore": 0,
        "capitalAfter": 0,
        "capitalIncreaseShares": 0,
        "shareholderSaleShares": 0,
        "extraSaleShares": 0,
        "fundUse": [],
        "highlights": [],
        "risks": ["Resmî izahname alanları henüz tam işlenmemiş olabilir."],
        "sources": [_source(source_url)],
        "agenda": [],
        "promises": [],
        "financials": [],
        "performance": None,
        "bulletinNo": "Bekleniyor",
        "approvalDate": "",
        "sourceUpdatedAt": now_iso,
    }
    item["status"] = compute_status(item)
    item["statusLabel"] = STATUS_LABELS[item["status"]]
    return item


def merge_public_records(
    items: list[dict[str, Any]],
    records: Iterable[dict[str, Any]],
    source_url: str,
    *,
    today: date | None = None,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    today = today or datetime.now(ZoneInfo("Europe/Istanbul")).date()
    now_iso = datetime.now(timezone.utc).isoformat()
    by_name = {normalize_company_name(item.get("company", "")): index for index, item in enumerate(items)}
    output = [deepcopy(item) for item in items]
    stats = {"matched": 0, "changed": 0, "added": 0, "ignored": 0}

    for record in records:
        key = normalize_company_name(record.get("company", ""))
        if not key:
            stats["ignored"] += 1
            continue
        index = _find_existing_index(key, by_name)
        if index is not None:
            merged, changed = _merge_record(output[index], record, source_url, now_iso)
            output[index] = merged
            by_name[key] = index
            stats["matched"] += 1
            stats["changed"] += int(changed)
        elif _should_add_new(record, today):
            item = _new_item(record, source_url, now_iso)
            by_name[key] = len(output)
            output.append(item)
            stats["added"] += 1
        else:
            stats["ignored"] += 1
    return output, stats
